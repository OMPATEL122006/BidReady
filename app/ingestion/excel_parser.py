import os
import re
from typing import Dict, List, Any
from app.config.logging import logger

class ExcelParser:
    """
    Parser for Excel (.xlsx, .xls) tender spreadsheets & BOQs.
    Trims trailing empty columns to prevent empty cell walls.
    """
    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse(self) -> Dict[int, List[Dict[str, Any]]]:
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        ext = os.path.splitext(self.file_path)[1].lower()
        if ext == ".xls":
            return self._parse_xls()
        else:
            return self._parse_xlsx()

    def _parse_xlsx(self) -> Dict[int, List[Dict[str, Any]]]:
        import openpyxl
        parsed_data = {}
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        
        for idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            raw_rows = []
            for row in sheet.iter_rows(values_only=True):
                str_row = [str(cell).strip() if cell is not None and str(cell).strip() != "None" else "" for cell in row]
                while str_row and str_row[-1] == "":
                    str_row.pop()
                if any(cell != "" for cell in str_row):
                    raw_rows.append(str_row)
                    
            if raw_rows:
                parsed_data[idx + 1] = self._process_sheet_rows(sheet_name, raw_rows)
                
        return parsed_data

    def _parse_xls(self) -> Dict[int, List[Dict[str, Any]]]:
        import xlrd
        parsed_data = {}
        wb = xlrd.open_workbook(self.file_path)
        
        for idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(idx)
            sheet_name = sheet.name
            raw_rows = []
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                str_row = []
                for cell in row:
                    if isinstance(cell, float) and cell.is_integer():
                        str_row.append(str(int(cell)))
                    else:
                        val = str(cell).strip()
                        str_row.append(val if val != "None" else "")
                        
                while str_row and str_row[-1] == "":
                    str_row.pop()
                if any(cell != "" for cell in str_row):
                    raw_rows.append(str_row)
                    
            if raw_rows:
                parsed_data[idx + 1] = self._process_sheet_rows(sheet_name, raw_rows)
                
        return parsed_data

    def _process_sheet_rows(self, sheet_name: str, sheet_rows: List[List[str]]) -> List[Dict[str, Any]]:
        elements = []
        is_boq = False
        headers_idx = 0
        boq_indices = {}

        boq_keywords = {
            "item_no": ["item no", "sl. no.", "sl no", "item number", "sr. no", "s.no"],
            "description": ["description", "item description", "particulars", "specification", "work description"],
            "quantity": ["quantity", "qty"],
            "unit": ["units", "unit"],
            "rate": ["estimated rate", "basic rate", "rate", "unit rate"],
            "amount": ["total amount", "amount"]
        }

        for r_idx in range(min(30, len(sheet_rows))):
            row_str = " ".join(sheet_rows[r_idx]).lower()
            matched_cols = 0
            temp_indices = {}
            
            for cat, kw_list in boq_keywords.items():
                for c_idx, cell_val in enumerate(sheet_rows[r_idx]):
                    cell_lower = cell_val.lower()
                    if any(kw in cell_lower for kw in kw_list):
                        temp_indices[cat] = c_idx
                        matched_cols += 1
                        break

            if matched_cols >= 3 and ("description" in temp_indices or "item_no" in temp_indices):
                is_boq = True
                headers_idx = r_idx
                boq_indices = temp_indices
                break

        headers = sheet_rows[headers_idx] if headers_idx < len(sheet_rows) else sheet_rows[0]
        valid_cols = [i for i, h in enumerate(headers) if str(h).strip() != ""]
        if not valid_cols:
            valid_cols = list(range(min(10, max(len(r) for r in sheet_rows))))

        clean_headers = [str(headers[i]).strip() for i in valid_cols if i < len(headers)]

        if is_boq:
            data_rows = sheet_rows[headers_idx + 1:]
            markdown_lines = []
            col_widths = [max(len(r[i]) if i < len(r) else 0 for r in sheet_rows[headers_idx:]) for i in valid_cols]
            col_widths = [max(w, 3) for w in col_widths]

            header_line = "| " + " | ".join(clean_headers[idx].ljust(col_widths[idx]) for idx in range(len(valid_cols))) + " |"
            sep_line = "| " + " | ".join("-" * col_widths[idx] for idx in range(len(valid_cols))) + " |"
            markdown_lines.append(header_line)
            markdown_lines.append(sep_line)

            structured_boq_rows = []
            for row in data_rows:
                row_cells = [row[i] if i < len(row) else "" for i in valid_cols]
                if any(c.strip() for c in row_cells):
                    row_line = "| " + " | ".join(row_cells[idx].ljust(col_widths[idx]) for idx in range(len(valid_cols))) + " |"
                    markdown_lines.append(row_line)

                item_val = row[boq_indices["item_no"]].strip() if "item_no" in boq_indices and boq_indices["item_no"] < len(row) else ""
                desc_val = row[boq_indices["description"]].strip() if "description" in boq_indices and boq_indices["description"] < len(row) else ""
                qty_val = row[boq_indices["quantity"]].strip() if "quantity" in boq_indices and boq_indices["quantity"] < len(row) else ""
                unit_val = row[boq_indices["unit"]].strip() if "unit" in boq_indices and boq_indices["unit"] < len(row) else ""
                rate_val = row[boq_indices["rate"]].strip() if "rate" in boq_indices and boq_indices["rate"] < len(row) else ""
                amount_val = row[boq_indices["amount"]].strip() if "amount" in boq_indices and boq_indices["amount"] < len(row) else ""

                item_match = re.match(r'^(\d+(?:\.\d+)*)', item_val)
                clean_item_no = item_match.group(1) if item_match else item_val

                def parse_num(v):
                    c = re.sub(r'[^\d.]', '', v)
                    try:
                        return float(c) if c else 0.0
                    except ValueError:
                        return 0.0

                if desc_val and not re.match(r'^\d+\.0$', desc_val):
                    structured_boq_rows.append({
                        "item_no": clean_item_no,
                        "description": desc_val,
                        "quantity": parse_num(qty_val),
                        "unit": unit_val,
                        "rate": parse_num(rate_val),
                        "amount": parse_num(amount_val)
                    })

            elements.append({
                "type": "boq",
                "text": "\n".join(markdown_lines),
                "sheet_name": sheet_name,
                "rows": structured_boq_rows
            })
        else:
            col_widths = [max(len(row[i]) if i < len(row) else 0 for row in sheet_rows) for i in valid_cols]
            col_widths = [max(w, 3) for w in col_widths]

            markdown_lines = []
            header_line = "| " + " | ".join(clean_headers[idx].ljust(col_widths[idx]) for idx in range(len(valid_cols))) + " |"
            sep_line = "| " + " | ".join("-" * col_widths[idx] for idx in range(len(valid_cols))) + " |"
            markdown_lines.append(header_line)
            markdown_lines.append(sep_line)

            for row in sheet_rows[1:]:
                row_cells = [row[i] if i < len(row) else "" for i in valid_cols]
                if any(c.strip() for c in row_cells):
                    row_line = "| " + " | ".join(row_cells[idx].ljust(col_widths[idx]) for idx in range(len(valid_cols))) + " |"
                    markdown_lines.append(row_line)

            elements.append({
                "type": "table",
                "text": "\n".join(markdown_lines),
                "sheet_name": sheet_name,
                "headers": clean_headers,
                "rows": sheet_rows[1:]
            })

        return elements
